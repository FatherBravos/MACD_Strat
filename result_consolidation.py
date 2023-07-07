import pandas as pd
import glob
import os
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Paths for each type of CSV
paths = [
    r"C:\Users\49176\Desktop\MACD_Strat\Test Results\Batch8_D_A_TSLA_Optimized\Batch8_CSVs\consistency_p_*.csv",
    r"C:\Users\49176\Desktop\MACD_Strat\Test Results\Batch8_D_A_TSLA_Optimized\Batch8_CSVs\expectancy_p_*.csv",
    r"C:\Users\49176\Desktop\MACD_Strat\Test Results\Batch8_D_A_TSLA_Optimized\Batch8_CSVs\res_*.csv",
    r"C:\Users\49176\Desktop\MACD_Strat\Test Results\Batch8_D_A_TSLA_Optimized\Batch8_CSVs\return_p_*.csv",
    r"C:\Users\49176\Desktop\MACD_Strat\Test Results\Batch8_D_A_TSLA_Optimized\Batch8_CSVs\risk_p_*.csv",
]

# Define the sheet names for the workbook
sheet_names = ['consistency', 'expectancy', 'res', 'return', 'risk']

# Remove the default sheet created and start with a fresh workbook
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Iterate over each path and type
for path, sheet_name in zip(paths, sheet_names):

    # Gather all CSV files in the directory of the same type
    files = glob.glob(path)

    # Create a list to store data
    data = []

    # Iterate over each file of the same type
    for f in files:
        # Read the CSV file, skip the first row (header), and transpose
        df = pd.read_csv(f, skiprows=1, header=None).T

        # Rename the columns to the first row of the DataFrame
        df.columns = df.iloc[0]

        # Drop the first row
        df = df[1:]

        # Add index column from file name
        index = os.path.basename(f).split('_')[-1].split('.')[0]
        df.insert(0, 'index', index)

        # Append DataFrame to list
        data.append(df)

    # Concatenate all dataframes in the list
    df_from_each_file = pd.concat(data, ignore_index=True)

    # Create a sheet for this DataFrame
    ws = wb.create_sheet(sheet_name)

    # Write the dataframe to a worksheet
    for r in df_from_each_file.to_records(index=False):
        ws.append(list(r))

# Save the Excel file
wb.save(r'C:\Users\49176\Desktop\MACD_Strat\Test Results\results.xlsx')
